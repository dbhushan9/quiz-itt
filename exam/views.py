from openpyxl import Workbook, load_workbook


from django.contrib.auth import authenticate, login, get_user_model
from django.http import HttpResponseRedirect,HttpResponse
from django.shortcuts import render, redirect
from django.urls import reverse, reverse_lazy
from django.views.generic.list import ListView
from django.views.generic.detail import DetailView
from django.views.generic.edit import DeleteView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from django.contrib.auth import authenticate, login
from django.contrib.auth import logout

from .models import Exam,Question,Choice,AnswerKey
from .forms import ExamForm
from .utils import generate_random_username

def home(request):
    context ={}
    
    if request.user.is_authenticated:
        if request.user.is_superuser:
            return redirect('exam:list')
        else:
            return redirect('submission:home')
    
    if request.POST:
        try:
            username = request.POST.get('username')
            password = request.POST.get('password')
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_superuser:
                    return redirect('exam:list')
                else:
                    return redirect('submission:home')
        except:
            context = {'username':'User does not Exist'}
        
    
    if request.POST:
      
        try:
            user = User.objects.get(username=username)
        except:
            print('user not found')
            context = {'username':'User does not Exist'}
        if 'user' not in context.keys():
            print('pass mismatch')
            context = {'password':'Password is Invalid'}
    print("context",context)
    return render(request,'base/home.html',context)

def homepage(request):
    user = request.user


@login_required
def logout_view(request):
    logout(request)
    return redirect('home')

class ExamListView(ListView):
    queryset = Exam.objects.all()
    template_name = 'exam/list.html'

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super(ExamListView, self).dispatch(*args, **kwargs)

@login_required
def CreateExamView(request,id=None):
    try:
        exam_obj = Exam.objects.get(id=id)
    except Exam.DoesNotExist:
        exam_obj = None
    print('dasdas')
    form = ExamForm(request.POST or None, instance=exam_obj)
    context = {
        'form' : form,
    }
    if request.POST:
        print('POST')
        if form.is_valid():
            print('form valid')
            exam_obj = form.save(commit=False)
            print(exam_obj.id)
            exam_obj.save()
            return redirect('exam:list')
    return render(request,'exam/create.html', context)

@login_required
def DeleteExamView(request,id):
    try:
        exam_obj = Exam.objects.get(id=id)
        exam_obj.delete()
    except Exam.DoesNotExist:
        pass
    print('successful delete')
    return redirect('exam:list')

@login_required
def QuestionListView(request,slug):
    exam_obj = Exam.objects.get(slug=slug)
    que_objs = Question.objects.filter(exam=exam_obj)
    print(que_objs)
    context = {
        'object_list':que_objs,
        'exam':exam_obj,
    }
    return render(request,'exam/qlist.html', context)


@login_required
def CreateMultipleQuestionView(request,slug):
    try:
        exam_obj = Exam.objects.get(slug=slug)
    except:
        return redirect('exam:list')
    context = {
        'exam': exam_obj
    }
    if request.POST:
        print('get')
        try:
            wb = load_workbook(request.FILES['excel_file'])
            sheet_ranges = wb['Sheet1']
            n=2; c1 = "A"+str(n); c2 = "B"+str(n); c3 = "C"+str(n); c4 = "D"+str(n); c5 = "E"+str(n); c6 = "F"+str(n);c7= "G"+str(n);
            while True:
                if sheet_ranges[c1].value:
                    try:
                        if sheet_ranges[c7].value:
                            question_obj = Question(description=str(sheet_ranges[c1].value),code=sheet_ranges[c7].value,exam=exam_obj); question_obj.save()
                        else:
                            question_obj = Question(description=str(sheet_ranges[c1].value),exam=exam_obj); question_obj.save()
                        choice_list = []
                        choice1 = Choice(description=str(sheet_ranges[c3].value),question=question_obj); choice1.save();
                        choice2 = Choice(description=str(sheet_ranges[c4].value),question=question_obj); choice2.save();
                        choice3 = Choice(description=str(sheet_ranges[c5].value),question=question_obj); choice3.save();
                        choice4 = Choice(description=str(sheet_ranges[c6].value),question=question_obj); choice4.save();
                        correct_pos = int(sheet_ranges[c2].value)
                        choice_list.extend([choice1, choice2, choice3, choice4])
                        
                        choice_correct = choice_list[correct_pos -1]
                        
                        answerKey_obj = AnswerKey(question=question_obj, choice=choice_correct); answerKey_obj.save()
                    except Exception as e:
                        print('exception2',e)
                else:
                    break
                n=n+1; c1 = "A"+str(n); c2 = "B"+str(n); c3 = "C"+str(n); c4 = "D"+str(n); c5 = "E"+str(n); c6 = "F"+str(n);c7 = "G"+str(n)
        except Exception as e:
            print('exception1',e)
            return render(request,'exam/excel_question.html',context)
        return redirect('exam:list')
    else:
        return render(request,'exam/excel_question.html',context)

@login_required
def question_excel(request):
    excel_data = [
        ['Question Description','correct opt number','option1','option2','option3','option4','Code']
    ]
    if excel_data:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Sheet1')
        for line in excel_data:
            ws.append(line)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Exam_Questions.xlsx'
    wb.save(response)
    return response

@login_required
def question_user(request):
    excel_data = [
        ['Username','Password']
    ]
    if excel_data:
        wb = Workbook(write_only=True)
        ws = wb.create_sheet('Sheet1')
        for line in excel_data:
            ws.append(line)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Quizit_users.xlsx'
    wb.save(response)
    return response



User = get_user_model()
@login_required
def get_users(request):

    if request.POST:
        try:
            wb = load_workbook(request.FILES['excel_file'])
            sheet_ranges = wb['Sheet1']
            n=2; c1 = "A"+str(n); c2 = "B"+str(n);
            while True:
                if sheet_ranges[c1].value:
                    try:
                        user_name =str(sheet_ranges[c1].value)
                        password = str(sheet_ranges[c2].value)
                        new_user = User.objects.create_user(user_name,'dummy@email.com',password)
                    except Exception as e:
                        print('exception2',e)
                else:
                    break
                n=n+1; c1 = "A"+str(n); c2 = "B"+str(n);
        except Exception as e:
            print('exception1',e)
            return render(request,'exam/excel_users.html')
        return redirect('exam:list')
    else:
        return render(request,'exam/excel_users.html')




    # wb = Workbook()
    # #ws = wb.create_sheet('Users')
    # sheet=wb.active
    # for i in range(5):
    #     user_name = generate_random_username(exam=str(exam_obj))
    #     #print( User.objects.make_random_password())
    #     password = User.objects.make_random_password()
    #     new_user = User.objects.create_user(user_name,'dummy@email.com',password)
    #     exam_obj.students.add(new_user)
    #     sheet.cell(row=i+1, column=1).value = user_name
    #     sheet.cell(row=i+1, column=2).value = password

    # response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    # response['Content-Disposition'] = f'attachment; filename={exam_obj}_users.xlsx'
    # wb.save(response)
    # return response